import io
import logging
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from telegram import Bot

logger = logging.getLogger(__name__)

# ── Color palette ─────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1D3557")   # Dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="457B9D")   # Medium blue
ACCENT_FILL   = PatternFill("solid", fgColor="F1FAEE")   # Very light mint (alt rows)
EXPENSE_FILL  = PatternFill("solid", fgColor="E63946")   # Red for expense total
INCOME_FILL   = PatternFill("solid", fgColor="2A9D8F")   # Teal for income total
BALANCE_POS   = PatternFill("solid", fgColor="52B788")   # Green — positive balance
BALANCE_NEG   = PatternFill("solid", fgColor="E63946")   # Red — negative balance
TITLE_FILL    = PatternFill("solid", fgColor="1D3557")   # Same as header

HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
TITLE_FONT    = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
SECTION_FONT  = Font(bold=True, color="1D3557", size=12, name="Calibri")
NORMAL_FONT   = Font(size=10, name="Calibri")
TOTAL_FONT    = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
BALANCE_FONT  = Font(bold=True, color="FFFFFF", size=13, name="Calibri")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin",   color="CCCCCC"),
    right=Side(style="thin",  color="CCCCCC"),
    top=Side(style="thin",    color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
THICK_BORDER = Border(
    left=Side(style="medium",   color="1D3557"),
    right=Side(style="medium",  color="1D3557"),
    top=Side(style="medium",    color="1D3557"),
    bottom=Side(style="medium", color="1D3557"),
)


def _auto_width(ws, min_width: int = 10, max_width: int = 45):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = min_width
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


def _style_header_row(ws, row: int, num_cols: int, fill=None):
    fill = fill or HEADER_FILL
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _write_section_title(ws, row: int, text: str, num_cols: int = 4):
    """Write a styled section title spanning num_cols columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = SECTION_FONT
    cell.alignment = LEFT
    ws.row_dimensions[row].height = 22


def _add_pie_chart(wb, ws_data, data_start_row: int, data_end_row: int,
                   title: str, anchor: str) -> PieChart:
    """Create a styled pie chart and anchor it to a worksheet cell."""
    chart = PieChart()
    chart.title = title
    chart.style  = 10      # built-in style — colourful slices
    chart.width  = 14      # cm
    chart.height = 12      # cm

    data_ref = Reference(ws_data, min_col=2, min_row=data_start_row - 1, max_row=data_end_row)
    cats_ref = Reference(ws_data, min_col=1, min_row=data_start_row,     max_row=data_end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Show percentage labels on slices
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showVal     = False
    chart.dataLabels.showCatName = True

    ws_data.add_chart(chart, anchor)
    return chart


def generate_excel_report(sql_result: dict, current_date: str | None = None) -> io.BytesIO:
    """
    Generates a professional Excel workbook from sql_result.
    sql_result must be: {"expenses": [...], "income": [...]}
    Returns an in-memory BytesIO buffer ready to send.
    """
    wb = openpyxl.Workbook()
    expenses = sql_result.get("expenses", [])
    income   = sql_result.get("income",   [])

    today_str = (current_date or datetime.now(timezone.utc).isoformat())[:10]

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 1 — Summary
    # ─────────────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.rightToLeft = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]
    title_cell.value = f"💰 Personal Finance Report — {today_str}"
    title_cell.font  = TITLE_FONT
    title_cell.fill  = TITLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20

    # ── Expenses by Category ──────────────────────────────────────────────────
    category_totals: dict[str, dict] = {}
    for row in expenses:
        cat = row.get("category", "Other")
        if cat not in category_totals:
            category_totals[cat] = {"total": 0.0, "count": 0}
        category_totals[cat]["total"] += float(row.get("amount") or 0)
        category_totals[cat]["count"] += 1

    _write_section_title(ws, 4, "📉 Expenses by Category")

    headers_exp = ["Category", "Total (EGP)", "Transactions", "Avg (EGP)"]
    for col, h in enumerate(headers_exp, 1):
        ws.cell(row=5, column=col, value=h)
    _style_header_row(ws, 5, 4)

    r = 6
    exp_data_start = r
    for cat, data in sorted(category_totals.items(), key=lambda x: -x[1]["total"]):
        avg = round(data["total"] / data["count"], 2) if data["count"] else 0
        values = [cat, round(data["total"], 2), data["count"], avg]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = THIN_BORDER
            cell.alignment = LEFT if c == 1 else RIGHT
            cell.font      = NORMAL_FONT
            if r % 2 == 0:
                cell.fill = ACCENT_FILL
        r += 1
    exp_data_end = r - 1

    # Expenses Grand Total row
    grand_total_expenses = sum(d["total"] for d in category_totals.values())
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        cell.fill   = EXPENSE_FILL
        cell.font   = TOTAL_FONT
        cell.border = THICK_BORDER
        cell.alignment = RIGHT
    ws.cell(row=r, column=1, value="TOTAL EXPENSES").alignment = LEFT
    ws.cell(row=r, column=2, value=round(grand_total_expenses, 2))
    ws.cell(row=r, column=3, value=len(expenses))
    r_exp_total = r
    r += 2

    # ── Income Summary ────────────────────────────────────────────────────────
    income_source_totals: dict[str, dict] = {}
    for row in income:
        source = row.get("item", "other")
        if source not in income_source_totals:
            income_source_totals[source] = {"total": 0.0, "count": 0}
        income_source_totals[source]["total"] += float(row.get("amount") or 0)
        income_source_totals[source]["count"] += 1

    _write_section_title(ws, r, "📈 Income by Source")
    r += 1

    headers_inc = ["Source", "Total (EGP)", "Transactions", "Avg (EGP)"]
    for col, h in enumerate(headers_inc, 1):
        ws.cell(row=r, column=col, value=h)
    _style_header_row(ws, r, 4, fill=SUBHEAD_FILL)
    r += 1

    inc_data_start = r
    for source, data in sorted(income_source_totals.items(), key=lambda x: -x[1]["total"]):
        avg = round(data["total"] / data["count"], 2) if data["count"] else 0
        values = [source, round(data["total"], 2), data["count"], avg]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border    = THIN_BORDER
            cell.alignment = LEFT if c == 1 else RIGHT
            cell.font      = NORMAL_FONT
            if r % 2 == 0:
                cell.fill = ACCENT_FILL
        r += 1
    inc_data_end = r - 1

    # Income Grand Total row
    grand_total_income = sum(d["total"] for d in income_source_totals.values())
    for c in range(1, 5):
        cell = ws.cell(row=r, column=c)
        cell.fill   = INCOME_FILL
        cell.font   = TOTAL_FONT
        cell.border = THICK_BORDER
        cell.alignment = RIGHT
    ws.cell(row=r, column=1, value="TOTAL INCOME").alignment = LEFT
    ws.cell(row=r, column=2, value=round(grand_total_income, 2))
    ws.cell(row=r, column=3, value=len(income))
    r += 2

    # ── NET BALANCE ───────────────────────────────────────────────────────────
    net = round(grand_total_income - grand_total_expenses, 2)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    balance_label = ws.cell(row=r, column=1,
                            value=f"{'✅' if net >= 0 else '⚠️'} NET BALANCE")
    balance_label.font      = BALANCE_FONT
    balance_label.fill      = BALANCE_POS if net >= 0 else BALANCE_NEG
    balance_label.alignment = CENTER
    balance_label.border    = THICK_BORDER

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    balance_val = ws.cell(row=r, column=3, value=net)
    balance_val.font      = BALANCE_FONT
    balance_val.fill      = BALANCE_POS if net >= 0 else BALANCE_NEG
    balance_val.alignment = CENTER
    balance_val.border    = THICK_BORDER
    ws.row_dimensions[r].height = 28

    # ─────────────────────────────────────────────────────────────────────────
    # Hidden "chart_data" sheet — stores data for pie charts
    # ─────────────────────────────────────────────────────────────────────────
    ws_cd = wb.create_sheet("_ChartData")
    ws_cd.sheet_state = "hidden"

    # Expenses chart data (column A–B)
    ws_cd.cell(row=1, column=1, value="Expense Category")
    ws_cd.cell(row=1, column=2, value="Amount (EGP)")
    exp_cd_start = 2
    for i, (cat, data) in enumerate(
        sorted(category_totals.items(), key=lambda x: -x[1]["total"]), start=exp_cd_start
    ):
        ws_cd.cell(row=i, column=1, value=cat)
        ws_cd.cell(row=i, column=2, value=round(data["total"], 2))
    exp_cd_end = ws_cd.max_row

    # Income chart data (column D–E)
    ws_cd.cell(row=1, column=4, value="Income Source")
    ws_cd.cell(row=1, column=5, value="Amount (EGP)")
    inc_cd_start = 2
    for i, (source, data) in enumerate(
        sorted(income_source_totals.items(), key=lambda x: -x[1]["total"]), start=inc_cd_start
    ):
        ws_cd.cell(row=i, column=4, value=source)
        ws_cd.cell(row=i, column=5, value=round(data["total"], 2))
    inc_cd_end = ws_cd.max_row

    # ── Pie Chart 1: Expenses by Category (embedded in Summary) ──────────────
    if category_totals and exp_cd_end >= exp_cd_start:
        chart_exp = PieChart()
        chart_exp.title  = "Expenses by Category"
        chart_exp.style  = 10
        chart_exp.width  = 15
        chart_exp.height = 13

        data_ref = Reference(ws_cd, min_col=2, min_row=1, max_row=exp_cd_end)
        cats_ref = Reference(ws_cd, min_col=1, min_row=2, max_row=exp_cd_end)
        chart_exp.add_data(data_ref, titles_from_data=True)
        chart_exp.set_categories(cats_ref)

        chart_exp.dataLabels = DataLabelList()
        chart_exp.dataLabels.showPercent = True
        chart_exp.dataLabels.showCatName = True
        chart_exp.dataLabels.showVal     = False

        ws.add_chart(chart_exp, "F2")    # Anchor: column F, row 2

    # ── Pie Chart 2: Income by Source (embedded in Summary) ──────────────────
    if income_source_totals and inc_cd_end >= inc_cd_start:
        chart_inc = PieChart()
        chart_inc.title  = "Income by Source"
        chart_inc.style  = 26
        chart_inc.width  = 15
        chart_inc.height = 13

        data_ref_i = Reference(ws_cd, min_col=5, min_row=1, max_row=inc_cd_end)
        cats_ref_i = Reference(ws_cd, min_col=4, min_row=2, max_row=inc_cd_end)
        chart_inc.add_data(data_ref_i, titles_from_data=True)
        chart_inc.set_categories(cats_ref_i)

        chart_inc.dataLabels = DataLabelList()
        chart_inc.dataLabels.showPercent = True
        chart_inc.dataLabels.showCatName = True
        chart_inc.dataLabels.showVal     = False

        ws.add_chart(chart_inc, "F24")   # Anchor: column F, row 24 (below expenses chart)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 2 — Expenses Transactions
    # ─────────────────────────────────────────────────────────────────────────
    if expenses:
        ws_exp = wb.create_sheet("Expenses")
        ws_exp.freeze_panes = "A3"

        ws_exp.merge_cells("A1:E1")
        ws_exp["A1"].value     = f"📉 Expenses Transactions — {today_str}"
        ws_exp["A1"].font      = TITLE_FONT
        ws_exp["A1"].fill      = TITLE_FILL
        ws_exp["A1"].alignment = CENTER
        ws_exp.row_dimensions[1].height = 28

        trans_headers = ["Date", "Item / Description", "Category", "Amount (EGP)", "Currency"]
        for col, h in enumerate(trans_headers, 1):
            ws_exp.cell(row=2, column=col, value=h)
        _style_header_row(ws_exp, 2, len(trans_headers))
        ws_exp.row_dimensions[2].height = 20

        for row_idx, row in enumerate(expenses, start=3):
            created_at = row.get("created_at")
            if hasattr(created_at, "strftime"):
                date_str = created_at.strftime("%Y-%m-%d %H:%M")
            else:
                date_str = str(created_at)[:16] if created_at else ""

            values = [
                date_str,
                row.get("item", ""),
                row.get("category", ""),
                float(row.get("amount") or 0),
                row.get("currency", "EGP"),
            ]
            for c, v in enumerate(values, 1):
                cell = ws_exp.cell(row=row_idx, column=c, value=v)
                cell.border    = THIN_BORDER
                cell.alignment = LEFT if c <= 3 else RIGHT
                cell.font      = NORMAL_FONT
                if row_idx % 2 == 0:
                    cell.fill = ACCENT_FILL

        # Totals footer
        total_row = len(expenses) + 3
        for c in range(1, 6):
            cell = ws_exp.cell(row=total_row, column=c)
            cell.fill   = EXPENSE_FILL
            cell.font   = TOTAL_FONT
            cell.border = THICK_BORDER
        ws_exp.cell(row=total_row, column=1, value="TOTAL").alignment = LEFT
        ws_exp.cell(row=total_row, column=4,
                    value=round(grand_total_expenses, 2)).alignment = RIGHT

        _auto_width(ws_exp)

    # ─────────────────────────────────────────────────────────────────────────
    # Sheet 3 — Income Transactions
    # ─────────────────────────────────────────────────────────────────────────
    if income:
        ws_inc = wb.create_sheet("Income")
        ws_inc.freeze_panes = "A3"

        ws_inc.merge_cells("A1:E1")
        ws_inc["A1"].value     = f"📈 Income Transactions — {today_str}"
        ws_inc["A1"].font      = TITLE_FONT
        ws_inc["A1"].fill      = PatternFill("solid", fgColor="2A9D8F")
        ws_inc["A1"].alignment = CENTER
        ws_inc.row_dimensions[1].height = 28

        inc_headers = ["Date", "Source Type", "Description", "Amount (EGP)", "Currency"]
        for col, h in enumerate(inc_headers, 1):
            ws_inc.cell(row=2, column=col, value=h)
        _style_header_row(ws_inc, 2, len(inc_headers), fill=SUBHEAD_FILL)
        ws_inc.row_dimensions[2].height = 20

        for row_idx, row in enumerate(income, start=3):
            created_at = row.get("created_at")
            if hasattr(created_at, "strftime"):
                date_str = created_at.strftime("%Y-%m-%d %H:%M")
            else:
                date_str = str(created_at)[:16] if created_at else ""

            values = [
                date_str,
                row.get("item", ""),
                row.get("description", "") or "",
                float(row.get("amount") or 0),
                row.get("currency", "EGP"),
            ]
            for c, v in enumerate(values, 1):
                cell = ws_inc.cell(row=row_idx, column=c, value=v)
                cell.border    = THIN_BORDER
                cell.alignment = LEFT if c <= 3 else RIGHT
                cell.font      = NORMAL_FONT
                if row_idx % 2 == 0:
                    cell.fill = ACCENT_FILL

        # Totals footer
        total_row = len(income) + 3
        for c in range(1, 6):
            cell = ws_inc.cell(row=total_row, column=c)
            cell.fill   = INCOME_FILL
            cell.font   = TOTAL_FONT
            cell.border = THICK_BORDER
        ws_inc.cell(row=total_row, column=1, value="TOTAL").alignment = LEFT
        ws_inc.cell(row=total_row, column=4,
                    value=round(grand_total_income, 2)).alignment = RIGHT

        _auto_width(ws_inc)

    # ── Serialize ─────────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def export_excel_report(
    sql_result: dict,
    bot: Bot,
    chat_id: int,
    current_date: str | None = None,
) -> bool:
    """
    Generates the Excel report and sends it via Telegram.
    Returns True on success, False on error.
    """
    try:
        buffer = generate_excel_report(sql_result, current_date)

        today_str = (current_date or datetime.now(timezone.utc).isoformat())[:10]
        filename  = f"finance_report_{today_str}.xlsx"

        n_exp = len(sql_result.get("expenses", []))
        n_inc = len(sql_result.get("income",   []))

        await bot.send_document(
            chat_id=chat_id,
            document=buffer,
            filename=filename,
            caption=(
                "📊 *تقريرك المالي الشامل جاهز!*\n"
                f"📉 {n_exp} مصروف   |   📈 {n_inc} دخل\n\n"
                "الملف يحتوي على:\n"
                "• ملخص المصاريف والدخل\n"
                "• Pie Charts لكل فئة 🥧\n"
                "• سجل كامل بكل المعاملات"
            ),
            parse_mode="Markdown",
        )
        return True

    except Exception as e:
        logger.error(f"Excel export error: {e}", exc_info=True)
        return False
